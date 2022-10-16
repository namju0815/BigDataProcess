#!/usr/bin/python3
from openpyxl import Workbook
import math

wb = Workbook()

ws = wb.active
ws.append(['id', 'name', 'midtern', 'final' , 'homework', 'attendance', 'total', 'grade'])
ws.append([20140001, 'Sophia', 23, 53, 41, 1])
ws.append([20140002, 'Emily', 94, 36, 33, 1])
ws.append([20140003, 'Lily', 37, 20, 46, 1])
ws.append([20140004, 'Olivia', 73, 100, 72, 1])
ws.append([20140005, 'Amelia', 93, 46, 0, 1])
ws.append([20150001, 'Isla', 6, 30, 58, 1])
ws.append([20150003, 'Isabella', 71, 51, 54, 1])
ws.append([20150005, 'Ava', 43, 62, 56, 1])
ws.append([20150007, 'Sophie', 48, 92, 14, 1])
ws.append([20150009, 'Chloe', 91, 64, 39, 1])

total=[]
for t in range(2,12):
	#total
	midtern = ws.cell(row = t, column = 3).value * 0.3
	final = ws.cell(row = t, column = 4).value * 0.35
	homework = ws.cell(row = t, column = 5).value * 0.34
	attendance = ws.cell(row = t, column = 6).value * 100 * 0.01
	total_value = midtern + final + homework + attendance
	total.append(total_value)
	ws.cell(row = t, column = 7, value = total_value)

total.sort(reverse=True)
dic = []

for i in range(len(total)):
	dic.append('Z')
	
Agrade = math.trunc(len(total)*0.3)
Bgrade = math.trunc(len(total)*0.7)
Cgrade = math.trunc(len(total)-Bgrade)

for index in range(len(total)):
	if(index < Agrade):
		dic[index] = 'A'
		for i in range(len(total)):
			if(total[index] == total[i]):
				dic[i] = dic[index]
	elif(index < Bgrade):
		dic[index] = 'B'
		for i in range(len(total)):
			if(total[index] == total[i]):
				dic[i] = dic[index]
	else:
		dic[index] = 'C'
		for i in range(len(total)):
			if(total[index] == total[i]):
				dic[i] = dic[index]
Acount=0
Bcount=0
Ccount=0
for index in range(len(total)):
	if(dic[index] == 'A'):
		Acount+=1
	if(dic[index] == 'B'):
		Bcount+=1
	if(dic[index] == 'C'):
		Ccount+=1
A_count = Acount//2
B_count = Bcount//2
C_count = Ccount//2

for index in range(0,A_count):
	dic[index] = 'A+'
	
for index in range(Acount,Acount+B_count):
	dic[index] = 'B+'
	
for index in range(Acount+Bcount,Acount+Bcount+C_count):
	dic[index] = 'C+'
	
for g in range(2,12):
	count = -1
	for t in total:
		count+=1	
		if(ws.cell(row = g, column = 7).value == t):
			ws.cell(row = g, column = 8, value = dic[count])
	
wb.save("student.xlsx")

